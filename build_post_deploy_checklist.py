"""
Build a post-deploy test checklist as an .xlsx so Ahmad can tick each
item + write Pass/Fail/Notes inline while testing.

Output: d:/GRQ vs. Cursor/WalaPlus_Post_Deploy_Test_Checklist.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = r"d:\GRQ vs. Cursor\WalaPlus_Post_Deploy_Test_Checklist.xlsx"

# ─── Source data ─────────────────────────────────────────────────────
# Each row: (section, area, test_step, expected_outcome)
SECTIONS = [
    {
        "title": "0. Smoke (do first — if anything here fails, STOP)",
        "color": "DC2626",  # red-600
        "rows": [
            ("Login", "Load login page", "qms-dashboard.replit.app/login loads without errors"),
            ("Login", "Favicon green", "Browser tab shows green shield (close + reopen tab if old gray icon persists — favicon cache is aggressive)"),
            ("Login", "Badge green", "Login page badge above 'WalaPlus' is green (was blue/indigo)"),
            ("Login", "Sign-in button green", "Sign in button is green (was blue/indigo)"),
            ("Login", "Sign-in works", "Sign in → land on the dashboard"),
        ],
    },
    {
        "title": "1. Branding sweep (visual only)",
        "color": "059669",  # emerald-600
        "rows": [
            ("Call Eval", "SDR Evaluation tab",
             "Form header, eval modal header, overall score card all green"),
            ("Guide page", "Welcome banner",
             "Banner is green (was blue/indigo)"),
            ("Projects page", "PMP table header",
             "Header is green"),
            ("QMS page", "Consultant promo tile",
             "Tile is green"),
            ("Scorecard page", "Avatar tile",
             "Tile is green"),
            ("Accept-invite page", "Header",
             "Header is green"),
        ],
    },
    {
        "title": "2.1 Call Evaluation — Call Records sub-tab",
        "color": "2563EB",  # blue-600
        "rows": [
            ("Call Records", "Table loads", "Table loads with all calls"),
            ("CRM Link", "Badge per row",
             "Each row shows Lead / Deal / CRM badge; click opens the right Zoho record (not just search)"),
            ("CRM Link", "Activity-fallback badge",
             "Calls linked via activity show yellow ★ badge (lower confidence)"),
            ("CRM Link", "Unlinked Link button",
             "Unlinked calls show 'Link' button next to Activity"),
            ("Analyze All Pending", "Dialog appears",
             "Click button → dialog shows actual count + cost estimate (was silent 'no calls' before)"),
            ("Analyze All Pending", "Progress + complete",
             "Accept → progress badge ticks 1/N → N/N → grid refreshes with calls moved from pending/processing to analyzed"),
        ],
    },
    {
        "title": "2.2 Call Evaluation — SDR Evaluation sub-tab",
        "color": "7C3AED",  # violet-600
        "rows": [
            ("Active Scorecard line", "Real name + version + count",
             "Top line shows actual scorecard name + version + attribute count (was stuck on 'Loading active scorecard…')"),
            ("Batch Evaluation panel", "Eligibility populates",
             "Click expand → eligible count + cost estimate populate"),
            ("Pick call → eval renders", "Render OK",
             "Pick any analyzed call from list → evaluation renders"),
            ("CRM linkage indicator", "Shown in eval header",
             "Lead/Deal badge OR 'Not linked — Search Zoho | Re-run matcher'"),
            ("🔊 Listen to call panel", "Audio loads + scrubs",
             "Player appears below Overall Score; loads, scrubs without re-buffering (HTTP Range working)"),
            ("Transcript browser", "Arabic renders",
             "Collapsible transcript renders Arabic correctly"),
            ("Manager Review buttons", "Three buttons visible",
             "Approve / Adjust / Disagree all visible in Manager Review panel"),
            ("Double-click Approve", "Single row only",
             "Double-click Approve → only ONE review row added (not two)"),
            ("Adjust mode modal", "Per-attribute editor + live recompute",
             "Adjust opens modal; live recompute of overall + dimensions; delta vs AI shown"),
            ("Adjust save updates score",
             "Score badge updates",
             "Adjust + save with note → call list badge updates with adjusted score; review history shows AI X → Adjusted Y (±Δ)"),
            ("Coaching Plan: Mark delivered button",
             "Button visible",
             "'Mark coaching delivered' button on Coaching Plan panel"),
            ("Coaching delivery modal", "Pre-fill + save",
             "Modal pre-fills with attributes + courses → save with commitment → panel shows 'COACHED <date>' badge"),
        ],
    },
    {
        "title": "2.3 Call Evaluation — Analytics sub-tab",
        "color": "DB2777",  # pink-600
        "rows": [
            ("Sentiment Distribution chart", "Real counts",
             "Real numbers (was hardcoded 60/30/10); tooltip 'N calls (X%)'"),
            ("QA Score Trends chart", "Real weeks",
             "Real week-date x-axis with n=count tooltips (was hardcoded Week 1-4 / 75-85)"),
            ("Agent Performance table", "Avg QA populated",
             "Avg QA Score column populated (not all '--')"),
            ("Coaching Sessions panel", "KPIs + sessions feed",
             "Total / Delivered / Rate / Hours / Δ cards populate; recent sessions feed shows entries; top-coached attributes populated"),
            ("AI Training Feedback panel", "Reviewed/Rate/Corrections",
             "Reviewed / Approval rate / Corrections show real numbers + top-corrected attributes list"),
        ],
    },
    {
        "title": "2.4 Call Evaluation — QMS Bridge sub-tab",
        "color": "0891B2",  # cyan-600
        "rows": [
            ("CRM scope description",
             "Updated text",
             "Reads 'Leads + Deals + activity fallback' (was 'Leads only')"),
        ],
    },
    {
        "title": "2.5 Call Evaluation — CRM Compliance sub-tab",
        "color": "0891B2",  # cyan-600
        "rows": [
            ("Compliance records",
             "Records appear",
             "Records now appear (or will populate once you run Analyze All Pending in 2.1)"),
        ],
    },
    {
        "title": "3. Batch API end-to-end (~5 min)",
        "color": "EA580C",  # orange-600
        "rows": [
            ("Submit batch",
             "Job row appears",
             "SDR Evaluation → Batch Evaluation panel → 'Submit for Batch Evaluation' → confirm → new row appears with status 'validating'"),
            ("Concurrent submit blocked",
             "Only one batch",
             "Click submit in a second tab simultaneously → should NOT create a second batch (advisory lock working)"),
            ("Batch completes",
             "Status progresses to completed",
             "Wait for status validating → in_progress → completed (or hit 'Sync now' to force poll)"),
            ("Calls show new evaluations",
             "Analytics updated",
             "After completion → calls show new SDR evaluations populated in Analytics"),
        ],
    },
    {
        "title": "4. Race-condition / hotfix verification",
        "color": "B91C1C",  # red-700
        "rows": [
            ("Review double-click",
             "Single row",
             "Approve/Disagree double-click doesn't create duplicate review rows (count stays at 1)"),
            ("Adjust unknown-dim attrs",
             "Still contributes to overall",
             "If you have a 4th-dimension scorecard, attributes outside people/process/governance still pull weight in overall"),
            ("Manual sync + Inngest race",
             "No duplicate audit",
             "Manual sync while Inngest poller running doesn't double-log sdr_batch_completed in event_logs"),
        ],
    },
    {
        "title": "5. Operational sanity",
        "color": "475569",  # slate-600
        "rows": [
            ("Network: no 404/500",
             "Clean",
             "Browser DevTools → Network → confirm no 404s or 500s on main page loads"),
            ("Console: no errors",
             "Clean",
             "DevTools → Console → no red errors on SDR Evaluation tab"),
        ],
    },
    {
        "title": "⚠️ Operational item — your side, not testable from the app",
        "color": "F59E0B",  # amber-500
        "rows": [
            ("Enable OpenAI auto-recharge",
             "platform.openai.com/billing",
             "Without it, every analysis silently fails if prepaid balance hits zero (like the $10-was-out incident before)"),
        ],
    },
]

# ─── Styling helpers ─────────────────────────────────────────────────
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

WHITE_BOLD = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
ROW_FONT = Font(name="Calibri", size=10)

# ─── Build workbook ──────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Post-Deploy Checklist"

# Top banner
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
banner = ws.cell(row=1, column=1, value="WalaPlus QMS — Post-Deploy Test Checklist")
banner.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
banner.alignment = Alignment(vertical="center", horizontal="center")
banner.fill = PatternFill("solid", fgColor="10B981")  # emerald-500
ws.row_dimensions[1].height = 32

# Sub-banner (date + instructions)
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
sub = ws.cell(
    row=2, column=1,
    value="Generated 2026-05-23 · Tick Status per row · Notes column for screenshots / repro / observations",
)
sub.font = Font(name="Calibri", size=10, italic=True, color="475569")
sub.alignment = Alignment(vertical="center", horizontal="center")
ws.row_dimensions[2].height = 20

# Column headers
HEADERS = ["#", "Area", "Test Step", "Expected Outcome", "Status", "Notes / Screenshot", "Severity if Failed"]
for col_idx, h in enumerate(HEADERS, start=1):
    c = ws.cell(row=3, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = PatternFill("solid", fgColor="334155")  # slate-700
    c.alignment = CENTER
    c.border = BORDER
ws.row_dimensions[3].height = 28

# Status dropdown values
status_dv = DataValidation(
    type="list", formula1='"Pass,Fail,Blocked,N/A,Not Run"', allow_blank=True
)
status_dv.error = "Pick from list"
status_dv.errorTitle = "Invalid status"
ws.add_data_validation(status_dv)

severity_dv = DataValidation(
    type="list", formula1='"P0 blocker,P1 must-fix,P2 should-fix,Cosmetic,—"', allow_blank=True
)
ws.add_data_validation(severity_dv)

# Write sections + rows
current_row = 4
test_num = 0
for section in SECTIONS:
    # Section header (merged row)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    sc = ws.cell(row=current_row, column=1, value=section["title"])
    sc.font = WHITE_BOLD
    sc.fill = PatternFill("solid", fgColor=section["color"])
    sc.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    for area, step, expected in section["rows"]:
        test_num += 1
        row = [test_num, area, step, expected, "", "", ""]
        for col_idx, val in enumerate(row, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=val)
            c.font = ROW_FONT
            c.alignment = WRAP
            c.border = BORDER
            if col_idx == 1:
                c.alignment = CENTER
                c.font = Font(name="Calibri", size=10, bold=True, color="64748B")
        # Attach dropdowns
        status_dv.add(f"E{current_row}")
        severity_dv.add(f"G{current_row}")
        # Subtle alternating row tint for readability
        if test_num % 2 == 0:
            for col_idx in range(1, 8):
                ws.cell(row=current_row, column=col_idx).fill = PatternFill(
                    "solid", fgColor="F8FAFC"
                )
        ws.row_dimensions[current_row].height = 38
        current_row += 1

# Column widths
widths = {
    "A": 5,   # #
    "B": 22,  # Area
    "C": 40,  # Test Step
    "D": 55,  # Expected Outcome
    "E": 12,  # Status
    "F": 38,  # Notes
    "G": 16,  # Severity
}
for col_letter, w in widths.items():
    ws.column_dimensions[col_letter].width = w

# Freeze header (rows 1-3)
ws.freeze_panes = "A4"

# ─── Summary tab ─────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
ws2["A1"] = "WalaPlus QMS — Test Run Summary"
ws2["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
ws2["A1"].fill = PatternFill("solid", fgColor="10B981")
ws2.merge_cells("A1:D1")
ws2["A1"].alignment = Alignment(vertical="center", horizontal="center")
ws2.row_dimensions[1].height = 28

summary_rows = [
    ("Total tests", f"={test_num}"),
    ("Passed", '=COUNTIF(\'Post-Deploy Checklist\'!E:E,"Pass")'),
    ("Failed", '=COUNTIF(\'Post-Deploy Checklist\'!E:E,"Fail")'),
    ("Blocked", '=COUNTIF(\'Post-Deploy Checklist\'!E:E,"Blocked")'),
    ("N/A", '=COUNTIF(\'Post-Deploy Checklist\'!E:E,"N/A")'),
    ("Not Run", '=COUNTIF(\'Post-Deploy Checklist\'!E:E,"Not Run")+(' + str(test_num) + '-COUNTA(\'Post-Deploy Checklist\'!E4:E1000))'),
    ("", ""),
    ("P0 blockers", '=COUNTIF(\'Post-Deploy Checklist\'!G:G,"P0 blocker")'),
    ("P1 must-fix", '=COUNTIF(\'Post-Deploy Checklist\'!G:G,"P1 must-fix")'),
    ("P2 should-fix", '=COUNTIF(\'Post-Deploy Checklist\'!G:G,"P2 should-fix")'),
    ("Cosmetic", '=COUNTIF(\'Post-Deploy Checklist\'!G:G,"Cosmetic")'),
]
for i, (label, formula) in enumerate(summary_rows, start=3):
    ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
    ws2.cell(row=i, column=2, value=formula)
    ws2.cell(row=i, column=1).alignment = Alignment(indent=1)

ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 16

# Save
wb.save(OUTPUT)
print(f"Wrote {test_num} tests across {len(SECTIONS)} sections → {OUTPUT}")
