"""
Generate the Duplicate Radar post-publish test checklist as an XLSX file
the user can comment on. One sheet, all sections, with dropdown validation
on the Status column and wide Comments + Screenshot columns.

Run: py scripts/gen_test_checklist_xlsx.py
Output: ../Duplicate_Radar_Test_Checklist_<date>.xlsx
"""
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ------------------------------------------------------------------
# Checklist content — sections + rows. Each row:
#   (test_id, what_to_check, expected_result)
# section header rows are tuples of length 1.
# ------------------------------------------------------------------

SECTIONS: list[tuple[str, list[tuple[str, str, str]]]] = [
    (
        "0. Smoke test (run first, 5 min)",
        [
            ("0.1", "Open /duplicates", "Dashboard loads, no JS errors in browser console (F12)"),
            ("0.2", "Click Sync Now", "Sticky progress bar appears at top, % advances, completes"),
            ("0.3", "Click any cluster row", "Cluster detail modal opens cleanly (records, action buttons, snapshots, action timeline sections all visible)"),
            ("0.4", "Click Export CSV (header button)", "File downloads as duplicate_radar_export_<ts>.csv, opens in Excel"),
            ("0.5", "Open the downloaded CSV", "5 new playbook columns at right edge: Recommended Action, Survivorship Rule, Owner to Consult, Why This Verdict, Due Date"),
            ("0.6", "Cluster modal: click Resolve + Verify on a test cluster", "Shows 'Verifying with Zoho…' spinner, then verified/failed alert"),
            ("0.7", "GET /api/duplicates/cross-module-overlaps in browser", "Returns valid JSON (not 500). Confirms R6 schema + endpoint healthy"),
        ],
    ),
    (
        "1. Original CRM data bugs (must verify these specific deals)",
        [
            ("1.1", "CS Lifecycle tab — Emdad Najed Company, مؤسسة شغف لحلول, UNITED COMPANY", "NONE should appear with phase_churn_desync Critical anymore (renewal-after-churn rule from PR #37)"),
            ("1.2", "CS Pipeline Overlap tab — same three deals", "If Termination + Renewal>Churn, verdict = BLOCK with reason re_engaged_renewal_after_churn (PR #45)"),
            ("1.3", "Domain column on CS Lifecycle for the three deals", "emdadnajed.com / thenizerksa.com / unitedinv.co should now populate (not '—'). If still empty, check violation message for diagnostic hint about Zoho API field name"),
            ("1.4", "Owner Accountability — Export on a previously-cancelled owner row", "Downloads cleanly (no more 'Cancelled · 17m ago' tray entries) — PR #44 shared-pool fix"),
        ],
    ),
    (
        "2. New features — R8: Go-to-cluster links",
        [
            ("2.1", "Lead Duplicates / Deal Duplicates / Contact Duplicates / Account Duplicates tabs", "Every Domain cell is a clickable blue link"),
            ("2.2", "Click any Domain cell in a record table", "Opens cluster detail modal directly (no copy-paste-search needed)"),
        ],
    ),
    (
        "3. New features — R9: Preflight plain-English reasons",
        [
            ("3.1", "Preflight Check tab → click Load sample → click Check", "Results table populates"),
            ("3.2", "Reason column on each result row", "Shows English explanations (not raw codes like 'active_phase:adoption')"),
            ("3.3", "Hover the verdict badge (BLOCK/REVIEW/etc.)", "Shows the same plain-English explanation as a tooltip"),
            ("3.4", "Hover the Reason text", "Shows the raw machine code as a tooltip (for debug discoverability)"),
        ],
    ),
    (
        "4. New features — R7: SSE scan progress + sticky bar",
        [
            ("4.1", "Click Sync Now", "Progress bar pinned at top of viewport, % advances smoothly without 2s jumps, module chips update without flicker"),
            ("4.2", "Scroll the cluster table while scan runs", "Progress bar stays pinned at top (sticky)"),
            ("4.3", "Refresh the tab mid-scan", "Progress bar re-attaches automatically and resumes showing progress"),
            ("4.4", "DevTools → Network tab during scan", "/scan-stream connection stays open (EventSource); /scan-status NOT polled every 2s"),
        ],
    ),
    (
        "5. New features — R4: Creation-rate trend chart",
        [
            ("5.1", "Open Executive Summary tab", "New 'Duplicate Creation Trend' card visible below the Source/Confidence row"),
            ("5.2", "Chart series", "Three lines: red (new duplicates), grey dashed (new records total), blue (duplicate rate %)"),
            ("5.3", "KPI cards on the right", "Three: Latest bucket count + delta vs previous, Latest dup rate %, Window total"),
            ("5.4", "Window selector — toggle 4 / 12 / 26 / 52 weeks", "Chart re-renders cleanly each time, no canvas overlap or memory leak"),
        ],
    ),
    (
        "6. New features — R1: Remediation Playbook columns",
        [
            ("6.1", "Trigger Export CSV (header button) and Export XLSX (header button)", "Both files download"),
            ("6.2", "CSV — 5 rightmost columns", "Recommended Action, Survivorship Rule, Owner to Consult, Why This Verdict, Due Date"),
            ("6.3", "XLSX — every type sheet (Leads / Deals / Contacts / Accounts / All Records)", "Same 5 columns appear at the right"),
            ("6.4", "Pick a cluster with primary = 'ACME Co'. Non-primary row's Recommended Action", "Reads: Merge into \"ACME Co\""),
        ],
    ),
    (
        "7. New features — R2: Per-owner Remediation Packet",
        [
            ("7.1", "Owner Accountability tab — each row's actions cell", "Two links visible: green 'Packet' + blue 'Export'"),
            ("7.2", "Click Packet for any owner", "Downloads duplicate-radar-packet-<owner>.xlsx"),
            ("7.3", "Open the XLSX", "4 sheets: Cover, Action Items, Raw Records, FAQ"),
            ("7.4", "Cover sheet", "Owner metrics match the dashboard row (total records, duplicate rate, RAG, etc.); 'Packet due by' date populated; dispute path + escalation contact listed"),
            ("7.5", "Action Items sheet", "5 playbook columns + identifying record info"),
            ("7.6", "FAQ sheet", "7 Q&A entries covering what's a duplicate, how to dispute, etc."),
        ],
    ),
    (
        "8. New features — R3: Resolve + Verify against Zoho",
        [
            ("8.1", "Test cluster you ALREADY merged in Zoho → Resolve + Verify", "Admin key prompt → 'Verifying with Zoho…' spinner → ✓ Verified — N record(s) confirmed deleted toast"),
            ("8.2", "Test cluster NOT YET merged in Zoho → Resolve + Verify", "⚠ Verification failed alert → snapshot viewer AUTO-OPENS (Follow-up 2)"),
            ("8.3", "Cluster status after failed verification", "Cluster flipped BACK to active in Clusters tab (didn't silently disappear)"),
            ("8.4", "Verification badge on a successfully-verified cluster", "Emerald badge in cluster modal showing 'verified' + notes"),
            ("8.5", "Verification badge on a failed-verification cluster", "Red badge showing 'failed' + notes explaining what's still in Zoho"),
        ],
    ),
    (
        "9. New features — R10: Pre-merge snapshots",
        [
            ("9.1", "Open any resolved cluster in modal", "'Pre-merge snapshots' section lists the snapshot from when you marked it resolved"),
            ("9.2", "Click View on a snapshot", "Snapshot viewer modal opens (z-60, on top of cluster modal); shows frozen cluster facts + records table"),
            ("9.3", "Snapshot-ordering fix verification (the bug we caught): Mark a cluster Resolved + Verify, pass a NEW primary_record_id (different from current)", "Open the resulting snapshot — primary should be the ORIGINAL, not your new choice"),
        ],
    ),
    (
        "10. New features — Follow-up 1: Action timeline",
        [
            ("10.1", "Any cluster modal", "'Action timeline' section visible below Pre-merge snapshots"),
            ("10.2", "Cluster with at least one Resolved action", "Row with green 'Resolved' chip, timestamp, performer, records-affected count"),
            ("10.3", "Cluster with an Ignored action", "Row with gray 'Ignored' chip"),
            ("10.4", "Cluster that's never had any action", "'No actions yet' placeholder"),
        ],
    ),
    (
        "11. New features — R6: Cross-Module overlaps tab",
        [
            ("11.1", "New 'Cross-Module' tab between Account Duplicates and CS Pipeline Overlap", "Visible with indigo dot; pending count badge shows on the tab if any cross-module clusters exist"),
            ("11.2", "Click the tab", "4 KPI cards populate: total, Lead↔Contact, Lead↔Account, ARR exposure"),
            ("11.3", "7 filter chips", "All / Lead↔Contact / Lead↔Account / Lead↔Deal / Contact↔Account / Contact↔Deal / Deal↔Account / 3+ modules — clicking each filters the table client-side"),
            ("11.4", "Click any cluster row (not a checkbox)", "Cluster detail modal opens showing per-record LINK/CLOSE/CONVERT recommendations"),
        ],
    ),
    (
        "12. New features — Follow-up 3: Bulk-close leads (DESTRUCTIVE — uses dry run first)",
        [
            ("12.1", "Cross-Module tab — rows with leads vs rows without", "Lead-bearing rows show a checkbox; rows with no Lead show '—' with tooltip explaining"),
            ("12.2", "Select 1+ rows", "Amber bulk action bar appears at top showing selected count"),
            ("12.3", "Click 'Dry run' button → admin key", "Result modal opens: 'Dry run — examined N clusters', per-cluster 'would-close' counts, NO actual Zoho writes (verify by checking a lead in Zoho — status unchanged)"),
            ("12.4", "Click 'Close lead records in Zoho' → admin key → confirm", "Result modal shows real closed counts; cross-module tab reloads with resolved clusters absent"),
            ("12.5", "Verify in Zoho UI", "The selected leads' Lead_Status = 'Lost Lead'; Description field contains the Duplicate Radar note"),
            ("12.6", "Re-run on the same clusters immediately", "Result modal shows leads_skipped > 0 (idempotency — already-Lost leads skipped)"),
            ("12.7", "Try selecting 30+ rows then execute", "Server clamps at 25; per_cluster array length ≤ 25 in result modal"),
        ],
    ),
    (
        "13. New features — R5: Preflight webhook (API test, curl)",
        [
            ("13.1", "curl -X POST .../api/duplicates/preflight/check with active CS customer domain", "verdict=block, should_create=false"),
            ("13.2", "Same curl with a never-seen domain", "verdict=pass, should_create=true"),
            ("13.3", "Same curl with empty body", "400 with 'at least one of domain/email/...'"),
            ("13.4", "Same curl without x-admin-key header", "401 unauthorized"),
        ],
    ),
    (
        "14. Environment / config verification (Replit DB shell)",
        [
            ("14.1", "DUPLICATE_RADAR_FIELD_COMPANY_DOMAIN env var", "Set if you discovered the correct Zoho API field name from the diagnostic; leave unset if the normalized fallback already works"),
            ("14.2", "DB: SELECT column_name FROM information_schema.columns WHERE table_name='duplicate_clusters' AND column_name IN ('verification_state','verification_at','verification_notes')", "Returns 3 rows (R3 migration applied)"),
            ("14.3", "DB: SELECT to_regclass('duplicate_cluster_snapshots')", "Returns the table name (not NULL) — R10 migration applied"),
            ("14.4", "DB: SELECT COUNT(*) FROM duplicate_clusters WHERE status='active'", "Returns a number (any value > 0 is fine, just verifying query runs)"),
        ],
    ),
    (
        "15. Bug-fix verifications (specific issues we caught)",
        [
            ("15.1", "Snapshot ordering (today's fix)", "Already covered in 9.3 — primary in snapshot = original, not operator's new choice"),
            ("15.2", "CSV null% guard (audit cleanup)", "Open exported CSV — Confidence column is empty for null values, not literally 'null%'"),
            ("15.3", "CS Overlap lifecycle_state custom phase (PR #41)", "If DUPLICATE_RADAR_CS_ACTIVE_PHASES is customized with a non-standard name, that phase shows as 'Active (custom phase)' not 'Renewal'. Skip if env not customized."),
        ],
    ),
    (
        "16. Failure-mode tests (optional but recommended)",
        [
            ("16.1", "Refresh dashboard while a scan is in flight", "Sticky progress bar re-attaches via SSE (doesn't sit hidden)"),
            ("16.2", "Click Export on 4 owner rows in quick succession", "All 4 download cleanly (PR #44 shared-pool fix prevents the previous 'cancelled' state)"),
            ("16.3", "Open /api/duplicates/snapshots/999999 in browser", "404 with clean error message, no 500"),
            ("16.4", "Bulk-close with 25+ selected — verify clamp", "Server response per_cluster array length ≤ 25"),
        ],
    ),
]


# ------------------------------------------------------------------
# Styling
# ------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F2937")     # gray-800
SECTION_FILL = PatternFill("solid", fgColor="DBEAFE")    # blue-100
ALT_ROW_FILL = PatternFill("solid", fgColor="F9FAFB")    # gray-50
WHITE = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
SECTION_FONT = Font(bold=True, name="Calibri", size=12, color="1E3A8A")  # blue-900
BODY_FONT = Font(name="Calibri", size=10)
WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="left")

THIN = Side(style="thin", color="D1D5DB")  # gray-300
ALL_BORDERS = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_checklist(out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Checklist"

    # Column setup: ID | Test | Expected | Status | Comments | Screenshot
    headers = [
        ("Section / ID", 18),
        ("What to check", 50),
        ("Expected result", 60),
        ("Status", 14),
        ("Comments", 50),
        ("Screenshot / link", 30),
    ]
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE
        cell.alignment = WRAP_CENTER
        cell.border = ALL_BORDERS
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 24

    # Freeze the header
    ws.freeze_panes = "A2"

    # Data-validation list for Status
    status_dv = DataValidation(
        type="list",
        formula1='"Pass,Fail,Blocked,Skip,Pending"',
        allow_blank=True,
        showDropDown=False,  # showDropDown=False actually MEANS show the arrow (openpyxl quirk)
    )
    status_dv.error = "Pick one of: Pass / Fail / Blocked / Skip / Pending"
    status_dv.errorTitle = "Invalid status"
    ws.add_data_validation(status_dv)

    row = 2
    for section_name, rows in SECTIONS:
        # Section header row spans all 6 columns
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        sec_cell = ws.cell(row=row, column=1, value=section_name)
        sec_cell.fill = SECTION_FILL
        sec_cell.font = SECTION_FONT
        sec_cell.alignment = WRAP_CENTER
        sec_cell.border = ALL_BORDERS
        ws.row_dimensions[row].height = 22
        row += 1

        for i, (test_id, what, expected) in enumerate(rows):
            is_alt = i % 2 == 1
            values = [test_id, what, expected, "", "", ""]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = BODY_FONT
                cell.alignment = WRAP_TOP
                cell.border = ALL_BORDERS
                if is_alt:
                    cell.fill = ALT_ROW_FILL
            # Apply data validation to the Status column for this row
            status_dv.add(f"D{row}")
            # Auto-size row height proportional to text length
            longest = max(len(what), len(expected))
            ws.row_dimensions[row].height = max(28, min(80, 14 + (longest // 40) * 14))
            row += 1

    # Summary tile in column G for at-a-glance progress
    summary_col = 8  # H — leave column G blank as a gutter
    ws.cell(row=1, column=summary_col, value="Summary").fill = HEADER_FILL
    ws.cell(row=1, column=summary_col).font = WHITE
    ws.cell(row=1, column=summary_col).alignment = WRAP_CENTER
    ws.column_dimensions[get_column_letter(summary_col)].width = 18
    ws.column_dimensions[get_column_letter(summary_col + 1)].width = 12

    labels = ["Pass", "Fail", "Blocked", "Skip", "Pending", "Untested"]
    for i, lbl in enumerate(labels):
        ws.cell(row=2 + i, column=summary_col, value=lbl).font = BODY_FONT
        if lbl == "Untested":
            # Empty cells in column D = untested
            formula = f'=COUNTBLANK(D2:D{row - 1})'
        else:
            formula = f'=COUNTIF(D2:D{row - 1},"{lbl}")'
        ws.cell(row=2 + i, column=summary_col + 1, value=formula).font = BODY_FONT

    # Total row
    ws.cell(row=2 + len(labels), column=summary_col, value="Total checks").font = Font(bold=True, name="Calibri", size=10)
    ws.cell(
        row=2 + len(labels),
        column=summary_col + 1,
        value=f"=SUM({get_column_letter(summary_col + 1)}2:{get_column_letter(summary_col + 1)}{1 + len(labels)})",
    ).font = Font(bold=True, name="Calibri", size=10)

    # Instructions in column J
    instr_col = 10
    ws.column_dimensions[get_column_letter(instr_col)].width = 70
    ws.cell(row=1, column=instr_col, value="How to use this checklist").fill = HEADER_FILL
    ws.cell(row=1, column=instr_col).font = WHITE
    ws.cell(row=1, column=instr_col).alignment = WRAP_CENTER

    instructions = [
        "1. Work top-down. Section 0 (Smoke test) is the gate — if any of 0.1–0.7 fail, stop and flag back.",
        "2. Status column (D) has a dropdown: Pass / Fail / Blocked / Skip / Pending.",
        "3. Put browser-console errors, HTTP response bodies, or any odd behaviour in Comments (E).",
        "4. For failed tests, paste a screenshot link (Google Drive / Slack) in column F.",
        "5. The Summary block in column H counts each status automatically.",
        "6. Don't run section 12 (Bulk-close leads) against production-critical clusters — use test data or run Dry run first.",
        "7. Section 13 (Preflight webhook) needs an admin key and the deployed URL — see commands in the chat.",
        "",
        "When done, send the file back with any failed/blocked rows highlighted in column E.",
    ]
    for i, line in enumerate(instructions, start=2):
        c = ws.cell(row=i, column=instr_col, value=line)
        c.font = BODY_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")

    # Save
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    out = (
        Path(__file__).resolve().parents[2]
        / f"Duplicate_Radar_Test_Checklist_{date.today().isoformat()}.xlsx"
    )
    write_checklist(out)
