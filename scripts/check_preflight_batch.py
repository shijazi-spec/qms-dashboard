#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Preflight batch self-check — run this on ANY preflight run BEFORE sending the
list to Sales. It auto-finds the newest FLAGGED + PASS Excel files in your
Downloads folder and reports, in plain language, whether the batch is safe:

  * the verdict breakdown (Block / Review / Duplicate / Pass)
  * the LEAK TEST — any company flagged as a client on some rows but PASSED on
    others (the failure mode that risked cold-calling existing clients)
  * the exact passed contacts to pull out if a leak is found

Run (one command):
    python "D:\\2_QMS Platform\\qms-dashboard\\scripts\\check_preflight_batch.py"

Or double-click  check_preflight_batch.bat  (same folder).

Optional: pass a folder or the two files explicitly:
    python check_preflight_batch.py "C:\\path\\to\\Downloads"
    python check_preflight_batch.py FLAGGED.xlsx PASS.xlsx
"""
import sys, os, glob
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    print("Missing 'openpyxl'. Install once with:  pip install openpyxl")
    sys.exit(2)

try:
    sys.stdout.reconfigure(encoding="utf-8")  # show Arabic / symbols on Windows
except Exception:
    pass


def downloads_dir():
    return os.path.join(os.path.expanduser("~"), "Downloads")


def newest(pattern, folder):
    files = glob.glob(os.path.join(folder, pattern))
    return max(files, key=os.path.getmtime) if files else None


def resolve_files(args):
    # Two explicit files given.
    xlsx = [a for a in args if a.lower().endswith(".xlsx")]
    if len(xlsx) == 2:
        f = next((x for x in xlsx if "flag" in os.path.basename(x).lower()), xlsx[0])
        p = next((x for x in xlsx if "pass" in os.path.basename(x).lower()), xlsx[1])
        return f, p
    folder = next((a for a in args if os.path.isdir(a)), downloads_dir())
    f = newest("*FLAGGED*.xlsx", folder) or newest("*flagged*.xlsx", folder)
    p = newest("*PASS*.xlsx", folder) or newest("*pass*.xlsx", folder)
    return f, p


def load_findings(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = "Findings" if "Findings" in wb.sheetnames else wb.sheetnames[-1]
    it = wb[sheet].iter_rows(values_only=True)
    hdr = [str(h) if h is not None else "" for h in next(it)]
    idx = {h: i for i, h in enumerate(hdr)}
    rows = [r for r in it if r and any(c is not None for c in r)]
    wb.close()
    return idx, rows


def cell(r, idx, name):
    i = idx.get(name)
    if i is None or i >= len(r):
        return ""
    v = r[i]
    return "" if v is None else str(v).strip()


def norm(s):
    return (s or "").strip().lower()


def main():
    f_path, p_path = resolve_files(sys.argv[1:])
    if not f_path or not p_path:
        print("Could not find both a FLAGGED and a PASS .xlsx in your Downloads.")
        print("Tip: pass them explicitly:  python check_preflight_batch.py FLAGGED.xlsx PASS.xlsx")
        sys.exit(2)

    print("Checking this batch:")
    print("  FLAGGED:", os.path.basename(f_path))
    print("  PASS   :", os.path.basename(p_path))
    print()

    fi, frows = load_findings(f_path)
    pi, prows = load_findings(p_path)

    # Verdict breakdown
    vc = Counter()
    for r in frows:
        vc[cell(r, fi, "Verdict").split("—")[0].strip() or "?"] += 1
    print("=== Verdict breakdown ===")
    for k, v in vc.most_common():
        print(f"   {v:5}  {k}")
    print(f"   {len(prows):5}  PASS (safe to import)")
    print()

    # LEAK TEST: a company flagged as an existing CLIENT (Block/Review) on some
    # rows but PASSED on others — the exact thing that must never reach Sales.
    flagged_client = defaultdict(set)
    for r in frows:
        vd = cell(r, fi, "Verdict").upper()
        if "BLOCK" in vd or "REVIEW" in vd:
            co = norm(cell(r, fi, "Company"))
            if co:
                flagged_client[co].add(cell(r, fi, "Company"))
    pass_rows_by_co = defaultdict(list)
    for r in prows:
        co = norm(cell(r, pi, "Company"))
        if co:
            pass_rows_by_co[co].append(r)

    leaks = sorted(set(flagged_client) & set(pass_rows_by_co))

    print("=== LEAK TEST (must be 0) ===")
    if not leaks:
        print("   0 leaks — no company is flagged as a client AND passed elsewhere.")
    else:
        print(f"   {len(leaks)} company(ies) flagged as a client but ALSO in PASS — pull these contacts:")
        for co in leaks:
            disp = next(iter(flagged_client[co]))
            print(f"\n   ! {disp}")
            for r in pass_rows_by_co[co]:
                print(f"       passed contact: {cell(r,pi,'Contact Name') or '-'} | {cell(r,pi,'Email') or '-'} | {cell(r,pi,'Phone') or '-'}")
    print()

    print("=== RESULT ===")
    if not leaks:
        print("   ✓ CLEAN — no known client leaked into PASS. Safe to send to Sales.")
        sys.exit(0)
    else:
        print(f"   ✗ HOLD — {len(leaks)} client(s) leaked into PASS (listed above). Remove those contacts")
        print("     from the PASS list before sending, and tell the team so the data can be fixed.")
        sys.exit(1)


if __name__ == "__main__":
    main()
